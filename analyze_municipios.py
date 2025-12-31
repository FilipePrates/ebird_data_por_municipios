#!/usr/bin/env python3
import argparse
import json
import os
import sys
import time
from datetime import datetime
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen


API_BASE = "https://api.ebird.org/v2"
REGION = "BR-RJ"
ENV_KEY = "E_BIRD_API_KEY"
ENV_REFRESH = "E_BIRD_REFRESH"
ENV_LOCALE = "E_BIRD_LOCALE"
ENV_CACHE_ONLY = "E_BIRD_CACHE_ONLY"
ENV_RARITY_CUTOFF = "E_BIRD_RARITY_CUTOFF"
ENV_USE_COMBINED = "E_BIRD_USE_COMBINED"
DATA_DIR = "data"
DEFAULT_LOCALE = "pt_BR"


def load_env_key():
    key = os.getenv(ENV_KEY)
    if key:
        return key
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if not os.path.exists(env_path):
        return None
    with open(env_path, "r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            name, value = line.split("=", 1)
            if name.strip() == ENV_KEY:
                return value.strip().strip('"').strip("'")
    return None


def get_json(path, api_key):
    url = f"{API_BASE}{path}"
    req = Request(url, headers={"X-eBirdApiToken": api_key})
    with urlopen(req, timeout=30) as resp:
        data = resp.read().decode("utf-8")
    return json.loads(data)


def get_county_regions(api_key):
    # Subnational2 returns municipios for a state code (BR-RJ).
    return get_json(f"/ref/region/list/subnational2/{REGION}", api_key)


def get_species_list(region_code, api_key):
    return get_json(f"/product/spplist/{region_code}", api_key)


def get_taxonomy(api_key, locale):
    path = f"/ref/taxonomy/ebird?fmt=json&locale={locale}"
    return get_json(path, api_key)


def format_table(rows, headers):
    widths = [len(h) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(str(cell)))
    line = "+".join("-" * (w + 2) for w in widths)
    def fmt_row(r):
        return "|".join(f" {str(r[i]).ljust(widths[i])} " for i in range(len(r)))
    out = [line, fmt_row(headers), line]
    out.extend(fmt_row(r) for r in rows)
    out.append(line)
    return "\n".join(out)


def write_xlsx(rows, headers, output_dir):
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:  # pragma: no cover - optional dependency
        print("Missing dependency: openpyxl (install with pip install openpyxl)", file=sys.stderr)
        return None

    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(output_dir, f"species_by_municipio_{REGION}_{ts}.xlsx")

    wb = Workbook()
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    header_font = Font(bold=True)
    align = Alignment(horizontal="left", vertical="center")

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align

    for row in rows:
        ws.append(row)

    # Adjust column widths for readability.
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Chart: Top 15 municipios by species count.
    chart = BarChart()
    chart.title = "Top 15 municipios por especies"
    chart.y_axis.title = "Especies"
    chart.x_axis.title = "Municipio"
    top_rows = min(15, len(rows))
    if top_rows > 0:
        data_ref = Reference(ws, min_col=3, min_row=1, max_row=top_rows + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=top_rows + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 12
        chart.width = 22
        ws.add_chart(chart, "E2")

    wb.save(out_path)
    return out_path


def load_cache(path):
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def save_cache(path, payload):
    tmp_path = f"{path}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)
    os.replace(tmp_path, path)


def build_taxonomy_map(taxonomy):
    by_code = {}
    for item in taxonomy:
        code = item.get("speciesCode")
        if not code:
            continue
        by_code[code] = {
            "scientific": item.get("sciName", ""),
            "common": item.get("comName", ""),
            "family": item.get("familyComName", ""),
            "order": item.get("order", ""),
            "category": item.get("category", ""),
        }
    return by_code


def build_species_counts(municipio_species):
    counts = {}
    for codes in municipio_species.values():
        for code in codes:
            counts[code] = counts.get(code, 0) + 1
    return counts


def get_rarity_set(species_counts, total_municipios, cutoff=None):
    if cutoff is None:
        cutoff = max(2, int(round(total_municipios * 0.05)))
    rare = {code for code, count in species_counts.items() if count <= cutoff}
    return rare, cutoff


def write_species_sheets(wb, municipio_species, taxonomy_map, header_fill, header_font, align):
    from openpyxl.styles import PatternFill

    ws = wb.create_sheet("Species_Overview")
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    headers = (
        "Species Code",
        "Nome comum",
        "Nome cientifico",
        "Familia",
        "Ordem",
        "Categoria",
        "Municipios",
        "Raridade",
    )
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align

    species_counts = build_species_counts(municipio_species)
    rare_set, rarity_cutoff = get_rarity_set(species_counts, len(municipio_species))
    print(
        f"Rarity cutoff <= {rarity_cutoff} municipios; {len(rare_set)} species raras.",
        file=sys.stderr,
    )

    rarity_fill = PatternFill("solid", fgColor="FFF2CC")
    for code, count in sorted(species_counts.items(), key=lambda x: x[1], reverse=True):
        info = taxonomy_map.get(code, {})
        rarity = f"Rara (<= {rarity_cutoff} municipios)" if code in rare_set else ""
        ws.append(
            (
                code,
                info.get("common", ""),
                info.get("scientific", ""),
                info.get("family", ""),
                info.get("order", ""),
                info.get("category", ""),
                count,
                rarity,
            )
        )
        if rarity:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = rarity_fill

    # Per-municipio sheets with species lists.
    for municipio, species_list in municipio_species.items():
        safe_title = municipio[:31] if municipio else "Municipio"
        mws = wb.create_sheet(safe_title)
        mws.page_setup.paperSize = mws.PAPERSIZE_A4
        mws.page_setup.orientation = mws.ORIENTATION_PORTRAIT
        mws.append(("Species Code", "Nome comum", "Nome cientifico", "Familia", "Ordem", "Categoria"))
        for col in range(1, 7):
            cell = mws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align
        for code in sorted(species_list):
            info = taxonomy_map.get(code, {})
            rarity = code in rare_set
            mws.append(
                (
                    code,
                    info.get("common", ""),
                    info.get("scientific", ""),
                    info.get("family", ""),
                    info.get("order", ""),
                    info.get("category", ""),
                )
            )
            if rarity:
                for col in range(1, 7):
                    mws.cell(row=mws.max_row, column=col).fill = rarity_fill
        for col_cells in mws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                value = cell.value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            mws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def add_report_sheet(wb, summary_rows):
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Report")
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    title = "Relatorio A4 - Riqueza de especies por municipio (RJ)"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    header_font = Font(bold=True)
    align = Alignment(horizontal="left", vertical="center")

    # Top 10 table
    ws["A3"] = "Top 10 municipios"
    ws["A3"].font = header_font
    ws.append(("Municipio", "Especies"))
    for col in range(1, 3):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
    for name, _code, count in summary_rows[:10]:
        ws.append((name, count))

    top_chart = BarChart()
    top_chart.title = "Top 10 municipios"
    top_chart.y_axis.title = "Especies"
    top_chart.x_axis.title = "Municipio"
    data_ref = Reference(ws, min_col=2, min_row=4, max_row=4 + min(10, len(summary_rows)))
    cats_ref = Reference(ws, min_col=1, min_row=5, max_row=4 + min(10, len(summary_rows)))
    top_chart.add_data(data_ref, titles_from_data=True)
    top_chart.set_categories(cats_ref)
    top_chart.height = 8
    top_chart.width = 16
    ws.add_chart(top_chart, "D3")

    # Bottom 10 table
    start_row = 16
    ws[f"A{start_row}"] = "Bottom 10 municipios"
    ws[f"A{start_row}"].font = header_font
    ws.append(("Municipio", "Especies"))
    for col in range(1, 3):
        cell = ws.cell(row=start_row + 1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
    bottom_rows = list(reversed(summary_rows))[:10]
    for name, _code, count in bottom_rows:
        ws.append((name, count))

    bottom_chart = BarChart()
    bottom_chart.title = "Bottom 10 municipios"
    bottom_chart.y_axis.title = "Especies"
    bottom_chart.x_axis.title = "Municipio"
    data_ref = Reference(
        ws,
        min_col=2,
        min_row=start_row + 1,
        max_row=start_row + 1 + min(10, len(bottom_rows)),
    )
    cats_ref = Reference(
        ws,
        min_col=1,
        min_row=start_row + 2,
        max_row=start_row + 1 + min(10, len(bottom_rows)),
    )
    bottom_chart.add_data(data_ref, titles_from_data=True)
    bottom_chart.set_categories(cats_ref)
    bottom_chart.height = 8
    bottom_chart.width = 16
    ws.add_chart(bottom_chart, "D16")

    # Histogram data
    counts = [row[2] for row in summary_rows]
    if counts:
        max_count = max(counts)
        bin_size = 50 if max_count > 200 else 25 if max_count > 100 else 10
        bins = list(range(0, max_count + bin_size, bin_size))
        hist_rows = []
        for i in range(len(bins) - 1):
            lo = bins[i]
            hi = bins[i + 1] - 1
            n = sum(1 for c in counts if lo <= c <= hi)
            hist_rows.append((f"{lo}-{hi}", n))

        hist_start = 28
        ws[f"A{hist_start}"] = "Distribuicao de riqueza"
        ws[f"A{hist_start}"].font = header_font
        ws.append(("Faixa", "Municipios"))
        for col in range(1, 3):
            cell = ws.cell(row=hist_start + 1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align
        for label, n in hist_rows:
            ws.append((label, n))

        hist_chart = BarChart()
        hist_chart.title = "Distribuicao de riqueza"
        hist_chart.y_axis.title = "Municipios"
        hist_chart.x_axis.title = "Especies"
        data_ref = Reference(
            ws,
            min_col=2,
            min_row=hist_start + 1,
            max_row=hist_start + 1 + len(hist_rows),
        )
        cats_ref = Reference(
            ws,
            min_col=1,
            min_row=hist_start + 2,
            max_row=hist_start + 1 + len(hist_rows),
        )
        hist_chart.add_data(data_ref, titles_from_data=True)
        hist_chart.set_categories(cats_ref)
        hist_chart.height = 8
        hist_chart.width = 16
        ws.add_chart(hist_chart, "D28")

    for idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def add_stats_sheet(wb, summary_rows, municipio_species, taxonomy_map):
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("Stats")
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    header_font = Font(bold=True)
    align = Alignment(horizontal="left", vertical="center")

    counts = [row[2] for row in summary_rows]
    counts_sorted = sorted(counts)
    total_municipios = len(counts)
    total_species = len({code for codes in municipio_species.values() for code in codes})
    zero_count = sum(1 for c in counts if c == 0)
    avg_count = sum(counts) / total_municipios if total_municipios else 0

    def percentile(p):
        if not counts_sorted:
            return 0
        k = int(round((p / 100) * (len(counts_sorted) - 1)))
        return counts_sorted[k]

    stats_rows = [
        ("Municipios (total)", total_municipios),
        ("Especies (unicas no estado)", total_species),
        ("Municipios com 0 especies", zero_count),
        ("Media de especies por municipio", round(avg_count, 1)),
        ("Mediana", percentile(50)),
        ("P25", percentile(25)),
        ("P75", percentile(75)),
        ("Maximo", counts_sorted[-1] if counts_sorted else 0),
        ("Minimo", counts_sorted[0] if counts_sorted else 0),
    ]

    ws["A1"] = "Resumo estatistico"
    ws["A1"].font = Font(bold=True, size=13)
    ws.append(("Indicador", "Valor"))
    for col in range(1, 3):
        cell = ws.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
    for row in stats_rows:
        ws.append(row)

    # Family summary (unique species by family).
    family_counts = {}
    for codes in municipio_species.values():
        for code in set(codes):
            fam = taxonomy_map.get(code, {}).get("family", "Unknown")
            family_counts[fam] = family_counts.get(fam, 0) + 1

    family_rows = sorted(family_counts.items(), key=lambda x: x[1], reverse=True)[:15]
    start_row = 14
    ws[f"A{start_row}"] = "Top 15 familias (especies unicas)"
    ws[f"A{start_row}"].font = header_font
    ws.append(("Familia", "Especies"))
    for col in range(1, 3):
        cell = ws.cell(row=start_row + 1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
    for fam, n in family_rows:
        ws.append((fam, n))

    fam_chart = BarChart()
    fam_chart.title = "Top 15 familias"
    fam_chart.y_axis.title = "Especies"
    fam_chart.x_axis.title = "Familia"
    data_ref = Reference(
        ws,
        min_col=2,
        min_row=start_row + 1,
        max_row=start_row + 1 + len(family_rows),
    )
    cats_ref = Reference(
        ws,
        min_col=1,
        min_row=start_row + 2,
        max_row=start_row + 1 + len(family_rows),
    )
    fam_chart.add_data(data_ref, titles_from_data=True)
    fam_chart.set_categories(cats_ref)
    fam_chart.height = 10
    fam_chart.width = 18
    ws.add_chart(fam_chart, "D14")


def add_clade_sheet(wb, municipio_species, taxonomy_map):
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("Clades")
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    header_font = Font(bold=True)
    align = Alignment(horizontal="left", vertical="center")

    # Global clade totals by order.
    order_counts = {}
    for codes in municipio_species.values():
        for code in set(codes):
            order = taxonomy_map.get(code, {}).get("order", "Unknown")
            order_counts[order] = order_counts.get(order, 0) + 1
    top_orders = sorted(order_counts.items(), key=lambda x: x[1], reverse=True)
    top_orders = [o for o, _n in top_orders[:12]]

    ws.append(("Municipio",) + tuple(top_orders))
    for col in range(1, len(top_orders) + 2):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align

    for municipio, codes in municipio_species.items():
        counts = {order: 0 for order in top_orders}
        for code in set(codes):
            order = taxonomy_map.get(code, {}).get("order", "Unknown")
            if order in counts:
                counts[order] += 1
        row = (municipio,) + tuple(counts[o] for o in top_orders)
        ws.append(row)

    # Chart: total species by order (top 12).
    chart_ws = wb.create_sheet("Clades_Summary")
    chart_ws.page_setup.paperSize = chart_ws.PAPERSIZE_A4
    chart_ws.page_setup.orientation = chart_ws.ORIENTATION_PORTRAIT
    chart_ws.append(("Ordem", "Especies unicas"))
    for col in range(1, 3):
        cell = chart_ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
    for order in top_orders:
        chart_ws.append((order, order_counts.get(order, 0)))

    clade_chart = BarChart()
    clade_chart.title = "Top 12 ordens (especies unicas)"
    clade_chart.y_axis.title = "Especies"
    clade_chart.x_axis.title = "Ordem"
    data_ref = Reference(chart_ws, min_col=2, min_row=1, max_row=1 + len(top_orders))
    cats_ref = Reference(chart_ws, min_col=1, min_row=2, max_row=1 + len(top_orders))
    clade_chart.add_data(data_ref, titles_from_data=True)
    clade_chart.set_categories(cats_ref)
    clade_chart.height = 10
    clade_chart.width = 18
    chart_ws.add_chart(clade_chart, "D2")


def add_rarities_sheet(wb, municipio_species, taxonomy_map):
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("Raridades")
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    header_font = Font(bold=True)
    align = Alignment(horizontal="left", vertical="center")
    rarity_fill = PatternFill("solid", fgColor="FFF2CC")

    headers = ("Species Code", "Nome comum", "Nome cientifico", "Familia", "Ordem", "Municipios")
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align

    species_counts = build_species_counts(municipio_species)
    rare_set, rarity_cutoff = get_rarity_set(species_counts, len(municipio_species))

    for code, count in sorted(species_counts.items(), key=lambda x: x[1]):
        if code not in rare_set:
            continue
        info = taxonomy_map.get(code, {})
        ws.append(
            (
                code,
                info.get("common", ""),
                info.get("scientific", ""),
                info.get("family", ""),
                info.get("order", ""),
                f"{count} (<= {rarity_cutoff})",
            )
        )
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).fill = rarity_fill


def add_municipio_ranking_sheet(wb, summary_rows):
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("Municipios_Ranking")
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    header_font = Font(bold=True)
    align = Alignment(horizontal="left", vertical="center")

    headers = ("Rank", "Municipio", "Codigo", "Especies", "Percentil", "Classe")
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align

    counts = [row[2] for row in summary_rows]
    counts_sorted = sorted(counts)
    n = len(counts_sorted)

    def percentile(value):
        if n == 0:
            return 0
        # Percentile rank based on <= value.
        rank = sum(1 for c in counts_sorted if c <= value)
        return round((rank / n) * 100)

    def class_label(pct):
        if pct >= 67:
            return "Alta"
        if pct >= 34:
            return "Media"
        return "Baixa"

    for idx, (name, code, count) in enumerate(summary_rows, start=1):
        pct = percentile(count)
        ws.append((idx, name, code, count, f"{pct}%", class_label(pct)))

    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def main():
    parser = argparse.ArgumentParser(description="eBird species by municipio for Rio de Janeiro.")
    parser.add_argument(
        "--refresh",
        action="store_true",
        help="Ignore cached JSON and re-download all data.",
    )
    parser.add_argument(
        "--locale",
        default=os.getenv(ENV_LOCALE, DEFAULT_LOCALE),
        help="eBird taxonomy locale (default: pt_BR).",
    )
    args = parser.parse_args()

    api_key = load_env_key()
    if not api_key:
        print(f"Missing {ENV_KEY} in environment or .env", file=sys.stderr)
        return 1

    refresh = args.refresh or os.getenv(ENV_REFRESH, "").strip() == "1"
    cache_only = os.getenv(ENV_CACHE_ONLY, "").strip() == "1"
    cutoff_env = os.getenv(ENV_RARITY_CUTOFF, "").strip()
    rarity_cutoff = int(cutoff_env) if cutoff_env.isdigit() else None
    use_combined = os.getenv(ENV_USE_COMBINED, "1").strip() != "0"

    try:
        print(f"Loading municipios for {REGION}...", file=sys.stderr)
        os.makedirs(DATA_DIR, exist_ok=True)
        counties_cache = os.path.join(DATA_DIR, f"municipios_{REGION}.json")
        counties = None if refresh else load_cache(counties_cache)
        if counties is None:
            if cache_only:
                print(f"Cache only enabled, missing {counties_cache}", file=sys.stderr)
                return 1
            counties = get_county_regions(api_key)
            save_cache(counties_cache, counties)
            print(f"Saved municipios cache to {counties_cache}", file=sys.stderr)
        else:
            print(f"Using cached municipios from {counties_cache}", file=sys.stderr)
    except (HTTPError, URLError) as exc:
        print(f"Failed to load municipios for {REGION}: {exc}", file=sys.stderr)
        return 1

    taxonomy_cache = os.path.join(DATA_DIR, f"taxonomy_{REGION}_{args.locale}.json")
    taxonomy = None if refresh else load_cache(taxonomy_cache)
    if taxonomy is None:
        try:
            if cache_only:
                print(f"Cache only enabled, missing {taxonomy_cache}", file=sys.stderr)
                return 1
            print(f"Loading taxonomy ({args.locale})...", file=sys.stderr)
            taxonomy = get_taxonomy(api_key, args.locale)
            save_cache(taxonomy_cache, taxonomy)
        except (HTTPError, URLError) as exc:
            print(f"Warning: failed to load taxonomy: {exc}", file=sys.stderr)
            taxonomy = []
    taxonomy_map = build_taxonomy_map(taxonomy)

    results = []
    municipio_species = {}
    combined_cache = os.path.join(DATA_DIR, f"municipio_species_{REGION}.json")
    combined = None
    if use_combined and not refresh:
        combined = load_cache(combined_cache)
        if combined:
            print(f"Using combined cache from {combined_cache}", file=sys.stderr)
            for item in combined:
                name = item.get("name")
                code = item.get("code")
                species = item.get("species", [])
                if not name or not code:
                    continue
                results.append((name, code, len(species)))
                municipio_species[name] = species

    if not results:
        total = len(counties)
        print(f"Found {total} municipios. Fetching species lists...", file=sys.stderr)
        combined = []
        for idx, county in enumerate(counties, start=1):
            code = county.get("code")
            name = county.get("name", code)
            if not code:
                continue
            try:
                print(f"[{idx}/{total}] {name} ({code})", file=sys.stderr)
                species_cache = os.path.join(DATA_DIR, f"species_{code}.json")
                species = None if refresh else load_cache(species_cache)
                if species is None:
                    if cache_only:
                        print(f"Cache only enabled, missing {species_cache}", file=sys.stderr)
                        continue
                    species = get_species_list(code, api_key)
                    save_cache(species_cache, species)
                else:
                    print(f"Using cached species list for {code}", file=sys.stderr)
            except (HTTPError, URLError) as exc:
                print(f"Warning: failed for {name} ({code}): {exc}", file=sys.stderr)
                continue
            results.append((name, code, len(species)))
            municipio_species[name] = species
            combined.append({"name": name, "code": code, "species": species})
            # Gentle pacing to avoid rate limiting.
            if idx < total:
                time.sleep(0.2)
        if combined:
            save_cache(combined_cache, combined)
            print(f"Saved combined cache to {combined_cache}", file=sys.stderr)

    results.sort(key=lambda r: r[2], reverse=True)
    headers = ("Municipio", "Code", "Species")
    print(format_table(results, headers))
    out_path = write_xlsx(results, headers, output_dir="outputs")
    if out_path and os.path.exists(out_path):
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            wb = load_workbook(out_path)
            header_fill = PatternFill("solid", fgColor="DDEBF7")
            header_font = Font(bold=True)
            align = Alignment(horizontal="left", vertical="center")
            write_species_sheets(wb, municipio_species, taxonomy_map, header_fill, header_font, align)
            add_report_sheet(wb, results)
            add_stats_sheet(wb, results, municipio_species, taxonomy_map)
            add_municipio_ranking_sheet(wb, results)
            add_clade_sheet(wb, municipio_species, taxonomy_map)
            if rarity_cutoff is not None:
                # Override rarity cutoff via env var for all sheets.
                species_counts = build_species_counts(municipio_species)
                rare_set, cutoff = get_rarity_set(species_counts, len(municipio_species), rarity_cutoff)
                print(
                    f"Using custom rarity cutoff <= {cutoff} municipios; {len(rare_set)} species raras.",
                    file=sys.stderr,
                )
            add_rarities_sheet(wb, municipio_species, taxonomy_map)
            wb.save(out_path)
        except Exception as exc:  # pragma: no cover - best effort enhancement
            print(f"Warning: failed to add species sheets: {exc}", file=sys.stderr)
    if out_path:
        print(f"Saved Excel report to {out_path}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
